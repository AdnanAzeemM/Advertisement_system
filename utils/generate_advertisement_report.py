import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.http import HttpResponse
from io import BytesIO

logger = logging.getLogger(__name__)


def generate_excel_report(data):
    wb = Workbook()
    ws = wb.active

    headers = ["ad_name", "start_date", "end_date", "location_name", "daily_visitors", "blocked"]
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=header)

    for i, record in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=record.ad.name)
        ws.cell(row=i, column=2, value=record.ad.start_date)
        ws.cell(row=i, column=3, value=record.ad.end_date)
        ws.cell(row=i, column=4, value=record.location.name)
        ws.cell(row=i, column=5, value=record.daily_visitors)
        ws.cell(row=i, column=6, value=record.blocked)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    tab = Table(displayName="Table1", ref=f"A1:F{i}")

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d')
    file_name = f'daily_visitors_{date_str}.xlsx'
    file_path = os.path.join(os.getcwd(), file_name)

    wb.save(file_path)
    print(f"File saved as: {file_path}")
    return file_path

